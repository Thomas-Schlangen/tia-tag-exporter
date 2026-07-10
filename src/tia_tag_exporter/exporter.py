"""Export von Tag-Daten in eine formatierte Excel-Datei."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

SHEET_NAMES: dict[str, str] = {
    "plc_tags": "PLC-Tags",
    "hmi_tags": "HMI-Tags",
    "db_variables": "DB-Variablen",
}

_DECKBLATT_FIELDS = ["Kunde", "Projekt", "Anlage", "Erstellt von", "Datum", "Version", "Bemerkung"]

_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_MAX_COLUMN_WIDTH = 60


class ExcelExporter:
    """Schreibt extrahierte Tag-Daten als formatierte Excel-Arbeitsmappe."""

    def export(self, data: dict[str, list[dict[str, Any]]], output_path: Path) -> None:
        """Erstellt eine Excel-Datei mit Deckblatt und einem Sheet pro Kategorie.

        Args:
            data: Mapping von Kategorie-Schlüssel (``plc_tags``, ``hmi_tags``,
                ``db_variables``) auf eine Liste von Records (Dicts mit gleichen Keys).
            output_path: Zielpfad der ``.xlsx``-Datei.
        """
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)

        workbook = Workbook()
        workbook.remove(workbook.active)

        self._write_deckblatt(workbook)

        for key, sheet_name in SHEET_NAMES.items():
            records = data.get(key)
            if not records:
                logger.debug("Keine Daten für Kategorie '%s', Sheet wird übersprungen", key)
                continue
            sheet = workbook.create_sheet(title=sheet_name)
            if key == "db_variables":
                self._write_db_variables_sheet(sheet, records)
            else:
                self._write_sheet(sheet, records)

        if len(workbook.sheetnames) == 1:
            logger.warning("Keine Daten zum Exportieren vorhanden — nur Deckblatt wird erzeugt")

        workbook.save(output_path)
        logger.info("Excel-Datei geschrieben: %s", output_path)

    @staticmethod
    def _write_deckblatt(workbook: Workbook) -> None:
        """Erstellt den ersten Tab "Deckblatt" mit auszufüllenden Projektfeldern."""
        sheet = workbook.create_sheet(title="Deckblatt", index=0)

        sheet.merge_cells("A1:B1")
        title_cell = sheet["A1"]
        title_cell.value = "TIA Tag Export"
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[1].height = 28

        for row_offset, field in enumerate(_DECKBLATT_FIELDS, start=3):
            label_cell = sheet.cell(row=row_offset, column=1, value=field)
            label_cell.font = Font(bold=True)
            label_cell.border = _THIN_BORDER

            value_cell = sheet.cell(row=row_offset, column=2, value="")
            value_cell.border = _THIN_BORDER

        sheet.column_dimensions["A"].width = 20
        sheet.column_dimensions["B"].width = 45

    @staticmethod
    def _write_row(sheet: Worksheet, row_index: int, values: list[Any]) -> None:
        """Schreibt eine Zeile zellenweise über ``sheet.cell()`` statt
        ``sheet.append()``.

        ``sheet.cell(row=, column=, value=)`` gibt die geschriebene Zelle
        direkt zurück (O(1)-Dict-Zugriff) — so lässt sich der
        Formel-Schutz (siehe unten) ohne zusätzlichen Row-Lookup anwenden.
        Eine frühere Version rief dafür nach ``sheet.append()`` zusätzlich
        ``sheet[row_index]`` auf; bei Sheets mit zehntausenden Zeilen
        (DB-Variablen) führte das live reproduzierbar zu einem
        Performance-Hänger (mehrere Minuten statt Sekunden für ~50.000
        Zeilen) und wurde deshalb durch dieses Vorgehen ersetzt.

        openpyxl interpretiert außerdem jeden String-Wert, der mit ``=``
        beginnt, automatisch als Excel-Formel (``cell.data_type`` wird
        ``"f"`` statt ``"s"``). Betrifft z. B. DB-Variablen-Kommentare aus
        TIA wie ``"=true if more than one error is present"`` — Excel
        würde beim Öffnen versuchen, das als Formel auszuwerten, statt den
        Text anzuzeigen (live gefunden). Deshalb wird der Datentyp für
        betroffene Zellen hier direkt auf String erzwungen.
        """
        for col_index, value in enumerate(values, start=1):
            cell = sheet.cell(row=row_index, column=col_index, value=value)
            if isinstance(value, str) and value.startswith("="):
                cell.data_type = "s"

    @staticmethod
    def _write_grouped_rows(
        sheet: Worksheet,
        headers: list[str],
        records: list[dict[str, Any]],
        row_builder: Callable[[dict[str, Any]], list[Any]],
        group_key_fn: Callable[[dict[str, Any]], Any],
        measure_content: bool,
    ) -> None:
        """Schreibt Kopfzeile + gruppierte, einklappbare Datenzeilen für ein Sheet.

        Gemeinsame Logik für PLC-Tags/HMI-Tags (``_write_sheet``) und
        DB-Variablen (``_write_db_variables_sheet``): Kopfzeile fett formatiert,
        Zeilen werden per ``group_key_fn`` gruppiert (Leerzeile beim
        Gruppenwechsel, ``outline_level`` fürs Einklappen), erste Zeile
        eingefroren, Spaltenbreiten automatisch angepasst.

        ``row_builder`` wandelt einen Record in die zu schreibende Werteliste
        um (in derselben Reihenfolge wie ``headers``); ``group_key_fn`` liefert
        den Gruppierungswert für einen Record.

        ``measure_content`` steuert, ob die Spaltenbreite zusätzlich am
        tatsächlichen Zellinhalt gemessen wird (über ``row_builder``, ein
        weiterer Durchlauf über alle Records — für PLC-/HMI-Tags mit wenigen
        hundert Zeilen unproblematisch) oder nur an der Kopfzeile (schneller,
        nötig für DB-Variablen mit teils zehntausenden Zeilen). Die
        Padding-Konstante unterscheidet sich entsprechend (siehe unten) — das
        ist Absicht, kein Bug: mit Inhaltsmessung braucht es weniger Puffer,
        ohne sie etwas mehr, damit auch ungemessene, tendenziell längere
        Werte nicht sofort abgeschnitten wirken.
        """
        ExcelExporter._write_row(sheet, 1, headers)

        header_font = Font(bold=True)
        for cell in sheet[1]:
            cell.font = header_font

        sheet.sheet_properties.outlinePr.summaryBelow = False

        current_group: Any = None
        row_index = 1
        for record in records:
            group_value = group_key_fn(record)
            if current_group is not None and group_value != current_group:
                row_index += 1  # Leerzeile zwischen Gruppen-Blöcken (Zelle bleibt ungeschrieben = leer)

            row_index += 1
            ExcelExporter._write_row(sheet, row_index, row_builder(record))

            if group_value == current_group:
                sheet.row_dimensions[row_index].outline_level = 1
            current_group = group_value

        sheet.freeze_panes = "A2"

        padding = 2 if measure_content else 4
        max_lengths = [len(str(header)) for header in headers]
        if measure_content:
            for record in records:
                for col_index, value in enumerate(row_builder(record)):
                    max_lengths[col_index] = max(max_lengths[col_index], len(str(value)))

        for col_index, max_length in enumerate(max_lengths, start=1):
            sheet.column_dimensions[get_column_letter(col_index)].width = min(
                max_length + padding, _MAX_COLUMN_WIDTH
            )

    @staticmethod
    def _write_sheet(sheet: Worksheet, records: list[dict[str, Any]]) -> None:
        """Schreibt PLC-Tags/HMI-Tags: Zeilen werden nach Spalte A
        ("Variablentabelle") gruppiert und lassen sich per ``outline_level``
        links über +/- ein- und ausklappen (wie beim DB-Variablen-Sheet).
        Zwischen den Tabellen-Blöcken steht je eine Leerzeile.
        """
        headers = list(records[0].keys())
        group_key = headers[0]

        ExcelExporter._write_grouped_rows(
            sheet,
            headers,
            records,
            row_builder=lambda record: [record.get(header, "") for header in headers],
            group_key_fn=lambda record: record.get(group_key),
            measure_content=True,
        )

    @staticmethod
    def _write_db_variables_sheet(sheet: Worksheet, records: list[dict[str, Any]]) -> None:
        """Schreibt das DB-Variablen-Sheet: Spalte A "DB-Name" (Gruppierungs-
        spalte), gefolgt von "Pfad" (Ordnerpfad als Text, Ebenen mit " - "
        verbunden), den Ordnerebenen als eigene Spalten, Variablenname und den
        übrigen Feldern. Zeilen desselben DBs werden per ``outline_level``
        gruppiert, sodass sie im DB links per +/- eingeklappt werden können.
        Zwischen den DB-Blöcken steht je eine Leerzeile.
        """
        max_depth = max((len(record.get("_folder_path", [])) for record in records), default=0)
        folder_headers = [f"Ordnerebene {i + 1}" for i in range(max_depth)]
        other_headers = [
            header for header in records[0].keys() if header not in ("Name", "_folder_path", "_db_name")
        ]
        headers = ["DB-Name", "Pfad", *folder_headers, "Variablenname", *other_headers]

        def _build_row(record: dict[str, Any]) -> list[Any]:
            folder_path = record.get("_folder_path", [])
            folder_cells = [folder_path[i] if i < len(folder_path) else "" for i in range(max_depth)]
            return [
                record.get("_db_name", ""),
                " - ".join(folder_path),
                *folder_cells,
                record.get("Name", ""),
                *[record.get(header, "") for header in other_headers],
            ]

        ExcelExporter._write_grouped_rows(
            sheet,
            headers,
            records,
            row_builder=_build_row,
            group_key_fn=lambda record: record.get("_db_name"),
            measure_content=False,
        )
